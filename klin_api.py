from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from wsgiref.simple_server import make_server
from html import escape
import re


def get_schedule(environ):
    url_args = environ['web_api.url_args']
    if url_args:
        url_args_dict = dict(user=escape(url_args[0]), date=escape(url_args[1]))
    else:
        return not_found()

    wb = load_workbook('C://temp//1//file1.xlsx', read_only=True)
    print(wb['SSM schedule'].title)

    target_column = 0
    for ColumnOfCellObjects in wb.active['A5':'Z5']:
        for CellObj in ColumnOfCellObjects:
            if str(CellObj.value) == url_args_dict['user']:
                target_column = column_index_from_string(str(CellObj.coordinate)[:1])
                print(target_column)
                break
    if target_column == 0:
        return [b'No user found']

    target_row = 0
    for RowOfCellObjects in wb.active['B1':'B5000']:
        for CellObj in RowOfCellObjects:
            if str(CellObj.value) == url_args_dict['date'] + ' 00:00:00':
                target_row = int(str(CellObj.coordinate)[1:])
                print(target_row)
                break

    if target_row == 0:
        return [b'No date found']

    return [url_args_dict['user'].encode() + b" on " + url_args_dict['date'].encode() + b' has shift ' + str(wb.active.cell(target_row, target_column).value).encode()]


def not_found():
    return [b'invalid request']

# regexp for a date extraction should be rewritten with a more specific
url_dispatcher = [
    (r'get/schedule/([a-zA-Z0-9 ]+)/([0-9]{4}-[0-9]{2}-[0-9]{2})/?$', get_schedule)
]


def web_api(environ, start_response):
    start_response('200 OK', [('Content-Type', 'text/html')])
    path = environ.get('PATH_INFO', '').lstrip('/')
    for regex, callback in url_dispatcher:
        match = re.search(regex, path)
        if match:
            environ['web_api.url_args'] = match.groups()
            return callback(environ)
        else:
            return not_found()


if __name__ == '__main__':
    http_srv = make_server('localhost', 55555, web_api)
    http_srv.serve_forever()